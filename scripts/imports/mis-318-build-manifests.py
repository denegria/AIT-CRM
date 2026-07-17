#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
READY_FILL = PatternFill("solid", fgColor="D9EAD3")
HOLD_FILL = PatternFill("solid", fgColor="F4CCCC")
GATE_FILL = PatternFill("solid", fgColor="FFF2CC")

LOCATIONS = {
    "Bound Brook": "bound-brook",
    "Plainfield": "plainfield",
}

CATALOG = [
    "Intro to English",
    "English 1",
    "English 2",
    "English 3",
    "English 4",
    "English 5",
    "English 6",
    "GED",
    "Citizenship Prep",
    "Computer",
    "Math",
]

REQUIRED_CATALOG_ADDITIONS = {"Computer", "Math"}


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def normalized_name(value) -> str:
    value = unicodedata.normalize("NFKD", text(value).lower())
    value = "".join(character for character in value if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", value)).strip()


def normalized_phone(value) -> str:
    digits = re.sub(r"\D", "", text(value))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits


def phone_options(value) -> list[str]:
    options = []
    for segment in re.split(r"\s*(?:/|;|\||\n)\s*", text(value)):
        digits = normalized_phone(segment)
        if 10 <= len(digits) <= 13 and digits not in options:
            options.append(digits)
    return options


def read_sheet(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [text(value) for value in next(values)]
        except StopIteration:
            return []
        if headers == ["note"]:
            notes = [text(row[0]) for row in values if row]
            if notes == ["No rows"]:
                return []
        return [dict(zip(headers, row)) for row in values]
    finally:
        workbook.close()


def edit_distance(left: str, right: str) -> int:
    previous = list(range(len(right) + 1))
    for index, left_char in enumerate(left, 1):
        current = [index]
        for right_index, right_char in enumerate(right, 1):
            current.append(min(
                current[-1] + 1,
                previous[right_index] + 1,
                previous[right_index - 1] + (left_char != right_char),
            ))
        previous = current
    return previous[-1]


def name_relation(left, right) -> tuple[str, float]:
    left_name = normalized_name(left)
    right_name = normalized_name(right)
    if not left_name or not right_name:
        return "missing", 0.0
    if left_name == right_name:
        return "exact", 1.0
    noise = {"computacion", "basica"}
    left_tokens = [token for token in left_name.split() if token not in noise]
    right_tokens = [token for token in right_name.split() if token not in noise]
    if not left_tokens or not right_tokens:
        return "missing", 0.0
    left_name = " ".join(left_tokens)
    right_name = " ".join(right_tokens)
    left_set = set(left_tokens)
    right_set = set(right_tokens)
    similarity = SequenceMatcher(None, left_name, right_name).ratio()
    if left_set == right_set:
        return "same_tokens_reordered", max(similarity, 0.98)
    if min(len(left_set), len(right_set)) >= 2 and (left_set <= right_set or right_set <= left_set):
        return "middle_or_extra_name_variant", max(similarity, 0.96)
    if min(len(left_tokens), len(right_tokens)) == 1:
        short_tokens, long_tokens = (left_tokens, right_tokens) if len(left_tokens) == 1 else (right_tokens, left_tokens)
        short = short_tokens[0]
        first = long_tokens[0]
        first_similarity = SequenceMatcher(None, short, first).ratio()
        if short in long_tokens[:2]:
            return "short_name_expansion", max(similarity, 0.95)
        if (
            short.startswith(first)
            or first.startswith(short)
            or first_similarity >= 0.80
            or (
                len(short) >= 4
                and len(first) >= 4
                and short[0] == first[0]
                and short[-1] == first[-1]
                and edit_distance(short, first) <= 2
            )
        ):
            return "short_name_spelling_variant", max(similarity, first_similarity)
    first_similarity = SequenceMatcher(None, left_tokens[0], right_tokens[0]).ratio()
    last_similarity = SequenceMatcher(None, left_tokens[-1], right_tokens[-1]).ratio()
    if last_similarity >= 0.80 and (
        first_similarity >= 0.80
        or (similarity >= 0.82 and first_similarity >= 0.65)
        or left_tokens[0].startswith(right_tokens[0])
        or right_tokens[0].startswith(left_tokens[0])
    ):
        return "likely_spelling_or_nickname_variant", max(similarity, first_similarity, last_similarity)
    if similarity >= 0.94 and left_tokens[0][0] == right_tokens[0][0] and last_similarity >= 0.84:
        return "minor_spelling_variant", similarity
    return "different_name", similarity


def map_course(value) -> tuple[str, str, str]:
    raw = text(value)
    normalized = normalized_name(raw)
    if not normalized:
        return "", "missing", "No trustworthy course/level in source"
    invalid_prefixes = ("teach", "presencial", "pago total", "lunes", "viernes")
    if normalized.startswith(invalid_prefixes):
        return "", "invalid_source_value", "Teacher/modality/payment/schedule metadata is not a course"
    if normalized in {"intro", "nivel intro", "introductory"}:
        return "Intro to English", "mapped_catalog", "Normalized introductory English label"
    if normalized == "ged":
        return "GED", "mapped_catalog", "Exact catalog course"
    if normalized in {"ciudadania", "citizenship"}:
        return "Citizenship Prep", "mapped_catalog", "Normalized citizenship label"
    if normalized in {"computer", "compu", "computacion", "computacion basic", "computacion basica", "ms word"}:
        return "Computer", "required_catalog_addition", "Owner approved Computer as a required controlled course"
    if normalized in {"matematicas", "math", "mathematics"}:
        return "Math", "required_catalog_addition", "Owner approved Math as a required controlled course"
    number_match = re.fullmatch(r"([1-6])(?: 0)?", normalized)
    if not number_match:
        number_match = re.match(r"(?:nivel|level)\s*([1-6])(?:\b|\s)", normalized)
    if number_match:
        return f"English {number_match.group(1)}", "mapped_catalog", "Normalized numeric English level"
    if normalized.startswith("libro azul"):
        return "", "manual_mapping_review", "Book/unit progress does not identify a trustworthy course level"
    return "", "manual_mapping_review", "Unrecognized historical course label"


def latest_lead(leads: list[dict]) -> dict:
    if not leads:
        return {}
    return max(leads, key=lambda row: text(row.get("updated_at")) or text(row.get("created_at")))


def contact_score(contact: dict, leads: list[dict]) -> tuple:
    latest = latest_lead(leads)
    status_rank = {
        "Enrolled": 6,
        "Follow Up": 5,
        "New Lead": 4,
        "Dropped / Quit": 3,
        "Retargeting": 2,
        "Not Interested": 1,
    }.get(text(latest.get("status")), 0)
    return (
        0 if contact.get("archived_at") else 1,
        status_rank,
        text(latest.get("updated_at")) or text(latest.get("created_at")),
        text(contact.get("updated_at")) or text(contact.get("created_at")),
    )


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, column in enumerate(sheet.columns, 1):
        width = min(max((len(text(cell.value)) for cell in column), default=0) + 2, 48)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)


def add_rows(workbook: Workbook, title: str, rows: list[dict], status_field: str = "") -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["note"])
        sheet.append(["No rows"])
        style_sheet(sheet)
        return
    headers = []
    seen = set()
    for row in rows:
        for key in row:
            if key not in seen:
                headers.append(key)
                seen.add(key)
    sheet.append(headers)
    for row in rows:
        sheet.append([text(row.get(header)) for header in headers])
    style_sheet(sheet)
    if status_field and status_field in headers:
        column = headers.index(status_field) + 1
        for row_number in range(2, sheet.max_row + 1):
            value = text(sheet.cell(row_number, column).value)
            if value.startswith(("insert", "create", "reuse", "merge", "skip")):
                fill = READY_FILL
            elif value.startswith(("hold", "defer", "exclude")):
                fill = HOLD_FILL
            else:
                fill = GATE_FILL
            sheet.cell(row_number, column).fill = fill


def save_workbook(path: Path, sheets: list[tuple[str, list[dict], str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows, status_field in sheets:
        add_rows(workbook, title, rows, status_field)
    workbook.save(path)
    workbook.close()


def split_ids(value) -> list[str]:
    return [part.strip() for part in text(value).split(";") if part.strip()]


def final_contact_actions(source_root: Path, snapshot: dict) -> tuple[list[dict], dict[str, dict]]:
    contacts_by_id = {text(row.get("id")): row for row in snapshot["contacts"]}
    leads_by_contact = defaultdict(list)
    for lead in snapshot["leads"]:
        leads_by_contact[text(lead.get("contact_id"))].append(lead)
    output = []
    by_identity = {}

    for location, slug in LOCATIONS.items():
        comparison = source_root / "approval/prod-readonly-comparison-20260716" / f"ait-usa-{slug}-inactive-prod-comparison-approval.xlsx"
        review = source_root / "approval/manual-contact-resolution-20260716" / f"ait-usa-{slug}-manual-contact-resolution-review.xlsx"
        base_rows = read_sheet(comparison, "Contact_Approval_Manifest")
        manual_rows = read_sheet(review, "All_Manual_Case_Evidence")
        manual_by_candidate = {text(row.get("candidate_id")): row for row in manual_rows}

        for base in base_rows:
            candidate_id = text(base.get("candidate_id"))
            identity_key = text(base.get("identity_key"))
            classification = text(base.get("classification"))
            manual = manual_by_candidate.get(candidate_id)
            target_id = ""
            duplicate_ids = []
            primary_phone = normalized_phone(base.get("phone"))
            historical_phones = []
            confidence = text(base.get("match_confidence"))
            resolution_reason = ""

            if classification == "exact_existing_contact":
                ids = split_ids(base.get("crm_contact_ids"))
                assert len(ids) == 1
                target_id = ids[0]
                final_action = "reuse_existing_contact_exact"
                resolution_reason = "Unique normalized name + phone production match"
            elif classification == "new_contact_candidate":
                final_action = "create_new_contact"
                resolution_reason = "No production Contact identity match"
            elif manual:
                resolution_action = text(manual.get("resolution_action"))
                target_id = text(manual.get("target_contact_id"))
                duplicate_ids = split_ids(manual.get("duplicate_contact_ids"))
                primary_phone = normalized_phone(manual.get("repaired_phone")) or primary_phone
                historical_phones = [normalized_phone(value) for value in split_ids(manual.get("historical_phone_options")) if normalized_phone(value)]
                confidence = text(manual.get("confidence"))
                resolution_reason = text(manual.get("resolution_reason"))
                if resolution_action == "defer_no_contact_information":
                    final_action = "defer_no_usable_identity"
                else:
                    final_action = resolution_action
            else:
                raise RuntimeError(f"Missing final manual resolution for {location} {candidate_id}")

            target = contacts_by_id.get(target_id, {})
            latest = latest_lead(leads_by_contact.get(target_id, [])) if target_id else {}
            latest_status = text(latest.get("status"))
            active_overlap = text(base.get("active_source_overlap")).lower() == "yes"
            if final_action == "defer_no_usable_identity":
                lifecycle_action = "none_deferred"
            elif active_overlap or (latest_status and latest_status != "Dropped / Quit"):
                lifecycle_action = "preserve_newer_or_active_lifecycle"
            elif final_action == "create_new_contact" or final_action.startswith("repair_phone_and_create") or final_action.startswith("create_"):
                lifecycle_action = "create_dropped_quit_lead_after_approval"
            else:
                lifecycle_action = "keep_or_set_dropped_quit_after_apply_recheck"

            planned_reference = target_id or (f"planned:{identity_key}" if final_action != "defer_no_usable_identity" else "")
            row = {
                "location": location,
                "candidate_id": candidate_id,
                "identity_key": identity_key,
                "student_name": text(base.get("student_name")),
                "primary_phone": primary_phone,
                "primary_phone_policy": "none_deferred" if final_action == "defer_no_usable_identity" else "inactive_workbook_authoritative",
                "phone_history_policy": "preserve_all_other_valid_numbers",
                "historical_phone_options": "; ".join(dict.fromkeys(phone for phone in historical_phones if phone and phone != primary_phone)),
                "final_contact_action": final_action,
                "target_contact_id": target_id,
                "duplicate_contact_ids": "; ".join(duplicate_ids),
                "planned_contact_reference": planned_reference,
                "confidence": confidence,
                "resolution_reason": resolution_reason,
                "active_roster_overlap": "yes" if active_overlap else "no",
                "current_crm_status": latest_status,
                "proposed_lifecycle_action": lifecycle_action,
                "lifecycle_guard": text(base.get("lifecycle_guard")),
                "source_cells": text(base.get("source_cells")),
                "data_apply_state": "held_pending_manifest_approval",
            }
            output.append(row)
            by_identity[identity_key] = row

    return output, by_identity


def final_inactive_course_actions(source_root: Path, contact_by_identity: dict[str, dict]) -> list[dict]:
    rows = []
    for location, slug in LOCATIONS.items():
        path = source_root / "approval/prod-readonly-comparison-20260716" / f"ait-usa-{slug}-inactive-prod-comparison-approval.xlsx"
        for row in read_sheet(path, "Course_Approval_Manifest"):
            rows.append({**row, "location": location})

    duplicate_groups = defaultdict(list)
    for index, row in enumerate(rows):
        mapped_course, _status, _note = map_course(row.get("raw_course_value"))
        if int(text(row.get("source_duplicate_count")) or 0) <= 1:
            continue
        key = (
            text(row.get("location")),
            text(row.get("identity_key")),
            mapped_course,
            text(row.get("start_date")),
            text(row.get("end_date")),
            text(row.get("outcome_reason")),
        )
        duplicate_groups[key].append(index)
    canonical_duplicate_index = {min(indices) for indices in duplicate_groups.values() if len(indices) > 1}
    noncanonical_duplicate_index = {index for indices in duplicate_groups.values() if len(indices) > 1 for index in indices if index != min(indices)}

    output = []
    for index, row in enumerate(rows):
        identity_key = text(row.get("identity_key"))
        contact = contact_by_identity.get(identity_key)
        mapped_course, mapping_status, mapping_note = map_course(row.get("raw_course_value"))
        source_ready = text(row.get("source_ready_for_course_match")).lower() == "yes"
        if not contact or contact["final_contact_action"] == "defer_no_usable_identity":
            action = "defer_contact_identity"
        elif not mapped_course or mapping_status in {"missing", "invalid_source_value", "manual_mapping_review"}:
            action = "defer_course_definition"
        elif not source_ready:
            action = "defer_source_course_review"
        elif index in noncanonical_duplicate_index:
            action = "skip_source_duplicate"
        elif text(row.get("existing_course_match")) == "exact_existing_course":
            action = "skip_exact_existing_course"
        elif text(row.get("existing_course_match")) == "same_course_name_review_dates":
            action = "hold_existing_course_date_review"
        elif mapped_course in REQUIRED_CATALOG_ADDITIONS:
            action = "insert_dropped_course_after_catalog_gate"
        else:
            action = "insert_dropped_course_after_approval"

        output.append({
            "location": text(row.get("location")),
            "source_sheet": text(row.get("source_sheet")),
            "source_cell": text(row.get("source_cell")),
            "student_name": text(row.get("student_name")),
            "identity_key": identity_key,
            "planned_contact_reference": contact.get("planned_contact_reference", "") if contact else "",
            "contact_action": contact.get("final_contact_action", "missing_contact_resolution") if contact else "missing_contact_resolution",
            "active_roster_overlap": contact.get("active_roster_overlap", "no") if contact else "no",
            "contact_lifecycle_guard": contact.get("lifecycle_guard", "Manual review") if contact else "Manual review",
            "raw_course_value": text(row.get("raw_course_value")),
            "mapped_course": mapped_course,
            "course_mapping_status": mapping_status,
            "course_mapping_note": mapping_note,
            "course_status": "dropped",
            "start_date": text(row.get("start_date")),
            "end_date": text(row.get("end_date")),
            "outcome_reason": text(row.get("outcome_reason")),
            "notes": text(row.get("course_notes")),
            "source_duplicate_role": "canonical" if index in canonical_duplicate_index else "duplicate" if index in noncanonical_duplicate_index else "unique",
            "proposed_course_action": action,
            "data_apply_state": "held_pending_product_gates_and_manifest_approval",
        })
    return output


def build_active_manifest(source_root: Path, snapshot: dict) -> tuple[list[dict], list[dict]]:
    contacts = snapshot["contacts"]
    contacts_by_id = {text(row.get("id")): row for row in contacts}
    leads_by_contact = defaultdict(list)
    for lead in snapshot["leads"]:
        leads_by_contact[text(lead.get("contact_id"))].append(lead)
    by_exact = defaultdict(list)
    by_phone = defaultdict(list)
    by_name = defaultdict(list)
    for contact in contacts:
        name = normalized_name(contact.get("name"))
        phone = normalized_phone(contact.get("phone"))
        if name and phone:
            by_exact[(name, phone)].append(contact)
        if phone:
            by_phone[phone].append(contact)
        if name:
            by_name[name].append(contact)

    source_rows = []
    for location, slug in LOCATIONS.items():
        path = source_root / "clean/buckets" / f"ait-usa-{slug}-active-students.xlsx"
        for row in read_sheet(path, "Roster_Source_Rows"):
            source_rows.append({**row, "manifest_location": location})

    identity_counts = Counter((normalized_name(row.get("student_name")), normalized_phone(row.get("phone_normalized") or row.get("phone_raw"))) for row in source_rows)
    unique_contacts = {}
    enrollment_rows = []

    for row in source_rows:
        name = normalized_name(row.get("student_name"))
        source_phone_options = phone_options(row.get("phone_raw"))
        if len(source_phone_options) > 1:
            phone = source_phone_options[-1]
            source_historical_phones = source_phone_options[:-1]
        else:
            phone = normalized_phone(row.get("phone_normalized") or row.get("phone_raw"))
            source_historical_phones = []
        identity = (name, phone)
        exact = by_exact.get(identity, []) if name and phone else []
        target = {}
        duplicate_ids = []
        historical_phones = list(source_historical_phones)
        relation = ""
        if not name or not 10 <= len(phone) <= 13:
            contact_action = "hold_missing_contact_identity"
        elif len(exact) == 1:
            contact_action = "reuse_existing_contact_exact"
            target = exact[0]
            relation = "exact"
        elif len(exact) > 1:
            ranked = sorted(exact, key=lambda contact: contact_score(contact, leads_by_contact.get(text(contact.get("id")), [])), reverse=True)
            target = ranked[0]
            duplicate_ids = [text(contact.get("id")) for contact in ranked[1:]]
            contact_action = "merge_exact_duplicate_contacts_then_reuse"
            relation = "exact"
        else:
            phone_matches = by_phone.get(phone, [])
            related = []
            for contact in phone_matches:
                candidate_relation, score = name_relation(row.get("student_name"), contact.get("name"))
                if candidate_relation not in {"different_name", "missing"}:
                    related.append((contact, candidate_relation, score))
            if len(related) == 1:
                target, relation, _score = related[0]
                contact_action = "reuse_existing_contact_name_variant"
            elif len(related) > 1:
                related.sort(key=lambda item: (item[2], contact_score(item[0], leads_by_contact.get(text(item[0].get("id")), []))), reverse=True)
                target, relation, _score = related[0]
                duplicate_ids = [text(item[0].get("id")) for item in related[1:]]
                contact_action = "hold_ambiguous_same_phone_name_variants"
            elif phone_matches:
                contact_action = "create_separate_contact_shared_phone"
                relation = "different_name"
            else:
                name_matches = by_name.get(name, [])
                if len(name_matches) == 1:
                    target = name_matches[0]
                    existing_phone = normalized_phone(target.get("phone"))
                    if existing_phone and existing_phone != phone:
                        historical_phones.append(existing_phone)
                    contact_action = "reuse_existing_contact_set_current_roster_phone_primary"
                    relation = "exact_name_current_roster_phone"
                elif len(name_matches) > 1:
                    contact_action = "hold_ambiguous_exact_name_contacts"
                    relation = "exact_name_multiple_contacts"
                else:
                    contact_action = "create_new_contact"
                    relation = "no_match"

        mapped_course, mapping_status, mapping_note = map_course(row.get("class_level_or_course"))
        ready = text(row.get("ready_for_active_import_dry_run")).lower() == "yes"
        target_id = text(target.get("id"))
        planned_reference = target_id or (f"planned:active:{name}|{phone}" if not contact_action.startswith("hold") else "")
        repeated_identity = identity_counts.get(identity, 0) > 1
        if not ready:
            enrollment_action = "hold_roster_lifecycle_status"
        elif contact_action.startswith("hold"):
            enrollment_action = "hold_contact_identity"
        elif not mapped_course:
            enrollment_action = "hold_course_definition"
        elif not text(row.get("start_date")):
            enrollment_action = "hold_missing_start_date"
        elif mapped_course in REQUIRED_CATALOG_ADDITIONS and repeated_identity:
            enrollment_action = "insert_active_enrollment_after_catalog_and_multi_active_gates"
        elif mapped_course in REQUIRED_CATALOG_ADDITIONS:
            enrollment_action = "insert_active_enrollment_after_catalog_gate"
        elif repeated_identity:
            enrollment_action = "insert_active_enrollment_after_multi_active_gate"
        else:
            enrollment_action = "insert_active_enrollment_after_approval"

        latest = latest_lead(leads_by_contact.get(target_id, [])) if target_id else {}
        enrollment = {
            "location": text(row.get("manifest_location")),
            "source_sheet": text(row.get("source_sheet")),
            "source_cell": text(row.get("source_cell")),
            "student_name": text(row.get("student_name")),
            "primary_phone": phone,
            "historical_phone_options": "; ".join(historical_phones),
            "contact_action": contact_action,
            "contact_match_relation": relation,
            "target_contact_id": target_id,
            "duplicate_contact_ids": "; ".join(duplicate_ids),
            "planned_contact_reference": planned_reference,
            "current_crm_status": text(latest.get("status")),
            "class_section_key": text(row.get("class_section_key")),
            "raw_course_value": text(row.get("class_level_or_course")),
            "mapped_course": mapped_course,
            "course_mapping_status": mapping_status,
            "course_mapping_note": mapping_note,
            "course_status": "active",
            "start_date": text(row.get("start_date")),
            "teacher": text(row.get("class_teacher")),
            "course_location": text(row.get("interpreted_location")),
            "modality": text(row.get("class_modality")),
            "class_time": text(row.get("student_schedule_override")) or text(row.get("class_time")),
            "class_days": text(row.get("class_days")),
            "scheduled_days_per_week": text(row.get("attendance_frequency")),
            "source_roster_status": text(row.get("source_roster_status")),
            "source_roster_status_basis": text(row.get("source_roster_status_basis")),
            "identity_has_multiple_current_enrollments": "yes" if repeated_identity else "no",
            "requires_multiple_active_enrollment_product_gate": "yes" if repeated_identity else "no",
            "proposed_enrollment_action": enrollment_action,
            "data_apply_state": "held_pending_inactive_lane_and_product_gates",
        }
        enrollment_rows.append(enrollment)

        unique_key = f"{name}|{phone}"
        existing = unique_contacts.get(unique_key)
        candidate = {
            "identity_key": unique_key,
            "student_name": text(row.get("student_name")),
            "primary_phone": phone,
            "historical_phone_options": "; ".join(historical_phones),
            "contact_action": contact_action,
            "target_contact_id": target_id,
            "duplicate_contact_ids": "; ".join(duplicate_ids),
            "planned_contact_reference": planned_reference,
            "enrollment_row_count": identity_counts.get(identity, 0),
            "locations": text(row.get("manifest_location")),
            "data_apply_state": "held_pending_inactive_lane_and_product_gates",
        }
        if not existing:
            unique_contacts[unique_key] = candidate
        elif text(row.get("manifest_location")) not in existing["locations"].split("; "):
            existing["locations"] = f"{existing['locations']}; {text(row.get('manifest_location'))}"

    return list(unique_contacts.values()), enrollment_rows


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def execution_manifest(path: Path, payload: dict) -> dict:
    payload = json.loads(json.dumps({**payload, "schemaVersion": 1, "approvalState": "held"}, default=text))
    payload["contentSha256"] = hashlib.sha256(canonical_json(payload).encode("utf-8")).hexdigest()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    path.chmod(0o600)
    return {
        "vaultPath": path.relative_to(path.parents[4]).as_posix(),
        "sizeBytes": path.stat().st_size,
        "sha256": sha256(path),
    }


def with_action_key(row: dict, action_key: str) -> dict:
    return {"idempotencyKey": action_key, **row}


def action_key(lane: str, kind: str, *parts) -> str:
    digest = hashlib.sha256("|".join(text(part) for part in parts).encode("utf-8")).hexdigest()[:32]
    return f"mis-318:{lane}:{kind}:{digest}"


def archive_file(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    destination.chmod(0o600)


def archive_sources(source_root: Path, vault_root: Path) -> list[dict]:
    file_map = [
        ("original", source_root / "originals/bound-brook-source.xlsx", vault_root / "sources/original/bound-brook-source.xlsx"),
        ("original", source_root / "originals/plainfield-source.xlsx", vault_root / "sources/original/plainfield-source.xlsx"),
        ("cleaned", source_root / "clean/buckets/ait-usa-bound-brook-active-students.xlsx", vault_root / "sources/cleaned/ait-usa-bound-brook-active-students.xlsx"),
        ("cleaned", source_root / "clean/buckets/ait-usa-plainfield-active-students.xlsx", vault_root / "sources/cleaned/ait-usa-plainfield-active-students.xlsx"),
        ("cleaned", source_root / "clean/buckets/ait-usa-bound-brook-inactive-students-cleaning.xlsx", vault_root / "sources/cleaned/ait-usa-bound-brook-inactive-students-cleaning.xlsx"),
        ("cleaned", source_root / "clean/buckets/ait-usa-plainfield-inactive-students-cleaning.xlsx", vault_root / "sources/cleaned/ait-usa-plainfield-inactive-students-cleaning.xlsx"),
        ("comparison", source_root / "approval/prod-readonly-comparison-20260716/ait-usa-bound-brook-inactive-prod-comparison-approval.xlsx", vault_root / "sources/evidence/ait-usa-bound-brook-inactive-prod-comparison-approval.xlsx"),
        ("comparison", source_root / "approval/prod-readonly-comparison-20260716/ait-usa-plainfield-inactive-prod-comparison-approval.xlsx", vault_root / "sources/evidence/ait-usa-plainfield-inactive-prod-comparison-approval.xlsx"),
        ("identity_resolution", source_root / "approval/manual-contact-resolution-20260716/ait-usa-bound-brook-manual-contact-resolution-review.xlsx", vault_root / "sources/evidence/ait-usa-bound-brook-manual-contact-resolution-review.xlsx"),
        ("identity_resolution", source_root / "approval/manual-contact-resolution-20260716/ait-usa-plainfield-manual-contact-resolution-review.xlsx", vault_root / "sources/evidence/ait-usa-plainfield-manual-contact-resolution-review.xlsx"),
        ("production_snapshot", source_root / "analysis/prod-crm-readonly-snapshot-20260716.json", vault_root / "sources/evidence/prod-crm-readonly-snapshot-20260716.json"),
        ("attendance_held", source_root / "clean/individual/ait-usa-bound-brook-active-attendance-history.xlsx", vault_root / "held/attendance/ait-usa-bound-brook-active-attendance-history-unvalidated.xlsx"),
        ("attendance_held", source_root / "clean/individual/ait-usa-plainfield-active-attendance-history.xlsx", vault_root / "held/attendance/ait-usa-plainfield-active-attendance-history-unvalidated.xlsx"),
    ]
    inventory = []
    for role, source, destination in file_map:
        if not source.exists():
            raise FileNotFoundError(source)
        archive_file(source, destination)
        inventory.append({
            "role": role,
            "vaultPath": destination.relative_to(vault_root.parent.parent).as_posix(),
            "sizeBytes": destination.stat().st_size,
            "sha256": sha256(destination),
        })
    return inventory


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--vault-root", type=Path, default=Path("private-imports/mis-318"))
    parser.add_argument("--index", type=Path, default=Path("docs/imports/mis-318/manifest-index.json"))
    args = parser.parse_args()
    source_root = args.source_root.resolve()
    vault_root = args.vault_root.resolve()
    index_path = args.index.resolve()

    snapshot = json.loads((source_root / "analysis/prod-crm-readonly-snapshot-20260716.json").read_text(encoding="utf-8"))
    if snapshot.get("mode") != "read_only":
        raise RuntimeError("Production snapshot is not marked read_only")
    fingerprint = snapshot.get("fingerprint", {})
    if fingerprint.get("neonBranchId") != "br-purple-bar-aphafrgp" or fingerprint.get("businessUnitName") != "AIT USA Institute":
        raise RuntimeError("Unexpected production snapshot scope")

    contact_actions, contact_by_identity = final_contact_actions(source_root, snapshot)
    inactive_courses = final_inactive_course_actions(source_root, contact_by_identity)
    active_contacts, active_enrollments = build_active_manifest(source_root, snapshot)

    manifest_dir = vault_root / "manifests"
    inactive_path = manifest_dir / "ait-usa-inactive-student-final-action-manifest.xlsx"
    active_path = manifest_dir / "ait-usa-active-enrollment-manifest.xlsx"

    inactive_contact_counts = Counter(row["final_contact_action"] for row in contact_actions)
    inactive_course_counts = Counter(row["proposed_course_action"] for row in inactive_courses)
    inactive_summary = [
        {"metric": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "lane", "value": "inactive_students_dropped_course_history"},
        {"metric": "data_apply_state", "value": "held"},
        {"metric": "production_snapshot_mode", "value": "read_only"},
        {"metric": "contact_candidates", "value": len(contact_actions)},
        {"metric": "historical_course_rows", "value": len(inactive_courses)},
        *({"metric": f"contact:{key}", "value": value} for key, value in sorted(inactive_contact_counts.items())),
        *({"metric": f"course:{key}", "value": value} for key, value in sorted(inactive_course_counts.items())),
    ]
    inactive_readme = [
        {"topic": "Purpose", "detail": "Final approval manifest for inactive students and historical Dropped / Quit course records."},
        {"topic": "Authority", "detail": "Read-only planning artifact. No Contact, lead, course, phone, lifecycle, schema, or production write occurred."},
        {"topic": "Phone recency", "detail": "Dated CRM/A2/source evidence wins; otherwise the last/rightmost original number is primary. Older valid numbers remain history."},
        {"topic": "Phone-history gate", "detail": "AIT CRM currently has only one Contact phone field. Historical valid numbers stay in this manifest until an alternate-phone/history model exists."},
        {"topic": "Active protection", "detail": "An inactive source row never downgrades a Contact also present in the corrected active roster or a newer CRM lifecycle."},
        {"topic": "Computer and Math", "detail": "Owner approved both as required controlled courses. Eligible rows are gated, not discarded."},
        {"topic": "Apply gate", "detail": "Hold until product slices, idempotent dry-run, manifest approval, and explicit production-write approval."},
    ]
    catalog_rows = [{
        "course_name": course,
        "current_catalog_state": "required_addition" if course in REQUIRED_CATALOG_ADDITIONS else "existing",
        "manifest_policy": "eligible_after_catalog_slice" if course in REQUIRED_CATALOG_ADDITIONS else "eligible",
    } for course in CATALOG]
    save_workbook(inactive_path, [
        ("Read_Me", inactive_readme, ""),
        ("Summary", inactive_summary, ""),
        ("Contact_Actions", contact_actions, "final_contact_action"),
        ("Duplicate_Merge_Actions", [row for row in contact_actions if row["duplicate_contact_ids"]], "final_contact_action"),
        ("Deferred_Contacts", [row for row in contact_actions if row["final_contact_action"].startswith("defer")], "final_contact_action"),
        ("Historical_Course_Actions", inactive_courses, "proposed_course_action"),
        ("Course_Review_Holds", [row for row in inactive_courses if row["proposed_course_action"].startswith(("defer", "hold"))], "proposed_course_action"),
        ("Required_Course_Catalog", catalog_rows, "current_catalog_state"),
    ])

    active_contact_counts = Counter(row["contact_action"] for row in active_contacts)
    active_enrollment_counts = Counter(row["proposed_enrollment_action"] for row in active_enrollments)
    active_summary = [
        {"metric": "generated_at", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "lane", "value": "active_student_enrollments"},
        {"metric": "data_apply_state", "value": "held_after_inactive_lane"},
        {"metric": "official_active_roster_rows", "value": len(active_enrollments)},
        {"metric": "unique_active_contact_identities", "value": len(active_contacts)},
        *({"metric": f"contact:{key}", "value": value} for key, value in sorted(active_contact_counts.items())),
        *({"metric": f"enrollment:{key}", "value": value} for key, value in sorted(active_enrollment_counts.items())),
    ]
    active_readme = [
        {"topic": "Purpose", "detail": "Approval manifest for the corrected 147-row active roster and current course enrollments."},
        {"topic": "Sequence", "detail": "Data is held until the inactive lane is finalized and approved. Attendance is not included."},
        {"topic": "Contact identity", "detail": "Exact name+phone reuse is preferred; clear unique same-phone variants may reuse; genuinely different names sharing phones remain separate Contacts."},
        {"topic": "Current phone", "detail": "The current active roster phone becomes primary when an exact-name Contact has an older different phone; the prior number remains history."},
        {"topic": "Course gates", "detail": "Computer requires a controlled catalog addition. Active classes need first-class section/schedule identity, and multiple simultaneous active enrollments require a product change because current validation allows only one active course per Contact."},
        {"topic": "Attendance", "detail": "Absolute last. Existing attendance extracts are unvalidated lineage only and must be regenerated against this final manifest."},
    ]
    save_workbook(active_path, [
        ("Read_Me", active_readme, ""),
        ("Summary", active_summary, ""),
        ("Unique_Contact_Actions", active_contacts, "contact_action"),
        ("Active_Enrollment_Actions", active_enrollments, "proposed_enrollment_action"),
        ("Enrollment_Holds", [row for row in active_enrollments if row["proposed_enrollment_action"].startswith("hold")], "proposed_enrollment_action"),
        ("Multiple_Active_Gate", [row for row in active_enrollments if row["requires_multiple_active_enrollment_product_gate"] == "yes"], "proposed_enrollment_action"),
        ("Required_Course_Catalog", catalog_rows, "current_catalog_state"),
    ])

    inactive_path.chmod(0o600)
    active_path.chmod(0o600)
    execution_dir = manifest_dir / "execution"
    generated_at = datetime.now(timezone.utc).isoformat()
    inactive_execution_path = execution_dir / "ait-usa-inactive-student-actions-v1.json"
    inactive_execution_inventory = execution_manifest(inactive_execution_path, {
        "manifestId": "mis-318-inactive-students-v1",
        "lane": "inactive",
        "generatedAt": generated_at,
        "sourceWorkbook": {
            "filename": inactive_path.name,
            "sha256": sha256(inactive_path),
        },
        "sourceProductionFingerprint": {
            "neonBranchId": fingerprint.get("neonBranchId"),
            "businessUnitName": fingerprint.get("businessUnitName"),
        },
        "sequence": {"afterLane": None, "attendanceSupported": False},
        "requiredProductGates": ["MIS-319", "MIS-320", "MIS-321", "MIS-322", "MIS-324"],
        "expectedCounts": {
            "contacts": len(contact_actions),
            "resolvedContacts": sum(1 for row in contact_actions if not row["final_contact_action"].startswith("defer")),
            "deferredContacts": sum(1 for row in contact_actions if row["final_contact_action"].startswith("defer")),
            "courses": len(inactive_courses),
            "actionableCourses": sum(1 for row in inactive_courses if row["proposed_course_action"].startswith("insert")),
        },
        "contactActions": [
            with_action_key(row, action_key("inactive", "contact", row.get("candidate_id"), row.get("identity_key")))
            for row in contact_actions
        ],
        "classSectionActions": [],
        "courseActions": [
            with_action_key(row, action_key(
                "inactive",
                "course",
                row.get("location"),
                row.get("source_sheet"),
                row.get("source_cell"),
                row.get("identity_key"),
                row.get("mapped_course"),
            ))
            for row in inactive_courses
        ],
    })
    inactive_content_sha = json.loads(inactive_execution_path.read_text(encoding="utf-8"))["contentSha256"]

    section_variants = defaultdict(set)
    for row in active_enrollments:
        section_key = text(row.get("class_section_key"))
        if not section_key:
            continue
        signature = (
            text(row.get("mapped_course")),
            text(row.get("teacher")),
            text(row.get("course_location")) or text(row.get("location")),
            text(row.get("modality")),
            text(row.get("class_time")),
            text(row.get("class_days")),
            text(row.get("scheduled_days_per_week")),
        )
        section_variants[section_key].add(signature)

    resolved_section_keys = {}
    for section_key, signatures in section_variants.items():
        for signature in signatures:
            resolved_section_keys[(section_key, signature)] = section_key if len(signatures) == 1 else (
                f"{section_key}-{hashlib.sha256(canonical_json(signature).encode('utf-8')).hexdigest()[:8]}"
            )

    active_sections = {}
    active_execution_rows = []
    for row in active_enrollments:
        original_section_key = text(row.get("class_section_key"))
        signature = (
            text(row.get("mapped_course")),
            text(row.get("teacher")),
            text(row.get("course_location")) or text(row.get("location")),
            text(row.get("modality")),
            text(row.get("class_time")),
            text(row.get("class_days")),
            text(row.get("scheduled_days_per_week")),
        )
        section_key = resolved_section_keys.get((original_section_key, signature), original_section_key)
        execution_row = {**row, "resolved_class_section_key": section_key}
        active_execution_rows.append(execution_row)
        if not section_key:
            continue
        section = {
            "sectionKey": section_key,
            "sourceSectionKey": original_section_key,
            "courseName": text(row.get("mapped_course")),
            "teacher": text(row.get("teacher")),
            "courseLocation": text(row.get("course_location")) or text(row.get("location")),
            "modality": text(row.get("modality")),
            "classTime": text(row.get("class_time")),
            "classDays": text(row.get("class_days")),
            "scheduledDaysPerWeek": text(row.get("scheduled_days_per_week")),
            "sourceType": "student_roster",
            "sourceReference": f"MIS-323:{text(row.get('source_sheet'))}:{text(row.get('source_cell'))}",
        }
        active_sections.setdefault(section_key, section)

    active_execution_path = execution_dir / "ait-usa-active-enrollment-actions-v1.json"
    active_execution_inventory = execution_manifest(active_execution_path, {
        "manifestId": "mis-318-active-enrollments-v1",
        "lane": "active",
        "generatedAt": generated_at,
        "sourceWorkbook": {
            "filename": active_path.name,
            "sha256": sha256(active_path),
        },
        "sourceProductionFingerprint": {
            "neonBranchId": fingerprint.get("neonBranchId"),
            "businessUnitName": fingerprint.get("businessUnitName"),
        },
        "sequence": {
            "afterLane": "inactive",
            "requiredPriorManifestSha256": inactive_content_sha,
            "attendanceSupported": False,
        },
        "requiredProductGates": ["MIS-319", "MIS-320", "MIS-321", "MIS-322", "MIS-324"],
        "expectedCounts": {
            "contacts": len(active_contacts),
            "enrollments": len(active_enrollments),
            "actionableEnrollments": sum(1 for row in active_enrollments if row["proposed_enrollment_action"].startswith("insert")),
            "heldEnrollments": sum(1 for row in active_enrollments if row["proposed_enrollment_action"].startswith("hold")),
            "classSections": len(active_sections),
        },
        "contactActions": [
            with_action_key(row, action_key("active", "contact", row.get("identity_key"), row.get("planned_contact_reference")))
            for row in active_contacts
        ],
        "classSectionActions": [
            with_action_key(section, action_key("active", "section", section_key))
            for section_key, section in sorted(active_sections.items())
        ],
        "courseActions": [
            with_action_key(row, action_key(
                "active",
                "enrollment",
                row.get("location"),
                row.get("source_sheet"),
                row.get("source_cell"),
                row.get("planned_contact_reference"),
                row.get("resolved_class_section_key"),
            ))
            for row in active_execution_rows
        ],
    })
    archived = archive_sources(source_root, vault_root)
    manifest_inventory = [
        {
            "role": "current_inactive_manifest",
            "vaultPath": inactive_path.relative_to(vault_root.parent.parent).as_posix(),
            "sizeBytes": inactive_path.stat().st_size,
            "sha256": sha256(inactive_path),
        },
        {
            "role": "current_active_manifest",
            "vaultPath": active_path.relative_to(vault_root.parent.parent).as_posix(),
            "sizeBytes": active_path.stat().st_size,
            "sha256": sha256(active_path),
        },
        {"role": "current_inactive_execution_manifest", **inactive_execution_inventory},
        {"role": "current_active_execution_manifest", **active_execution_inventory},
    ]
    index = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "mode": "read_only_artifact_build",
        "productionFingerprint": {
            "neonBranchId": fingerprint.get("neonBranchId"),
            "businessUnitName": fingerprint.get("businessUnitName"),
        },
        "lanes": {
            "inactive": {
                "state": "manifest_final_data_held",
                "contactCandidates": len(contact_actions),
                "resolvedContactActions": sum(1 for row in contact_actions if not row["final_contact_action"].startswith("defer")),
                "deferredContacts": sum(1 for row in contact_actions if row["final_contact_action"].startswith("defer")),
                "historicalCourseRows": len(inactive_courses),
                "actionableCourseRowsAfterGates": sum(1 for row in inactive_courses if row["proposed_course_action"].startswith("insert")),
            },
            "active": {
                "state": "manifest_built_data_held_after_inactive",
                "officialRosterRows": len(active_enrollments),
                "uniqueContactIdentities": len(active_contacts),
                "actionableEnrollmentRowsAfterGates": sum(1 for row in active_enrollments if row["proposed_enrollment_action"].startswith("insert")),
                "heldEnrollmentRows": sum(1 for row in active_enrollments if row["proposed_enrollment_action"].startswith("hold")),
            },
            "attendance": {
                "state": "absolute_last_unvalidated_lineage_only",
                "approvedManifest": False,
                "requiredGate": "regenerate_and_revalidate_against_final_active_manifest_and_original_workbooks",
            },
        },
        "linearIssues": {
            "inactiveManifestLane": "MIS-318",
            "activeEnrollmentLane": "MIS-323",
            "attendanceLane": "MIS-272",
            "alternatePhoneHistoryGate": "MIS-319",
            "courseCatalogGate": "MIS-320",
            "classSectionsAndMultiEnrollmentGate": "MIS-321",
            "manifestImportServiceGate": "MIS-322",
            "safeContactMergeGate": "MIS-324",
        },
        "requiredProductGates": [
            "Add Computer and Math to the controlled AIT USA course catalog",
            "Add a Contact alternate-phone/history model so prior valid numbers are preserved without overloading contact_people",
            "Add first-class course-section identity and support multiple simultaneous active enrollments for one Contact",
            "Preserve class schedule, modality, teacher, location, and source lineage on imported active enrollments",
            "Provide an idempotent approval-gated manifest dry-run/apply path",
            "Provide collision-safe duplicate Contact relationship reparenting before merge actions apply",
        ],
        "files": sorted(archived + manifest_inventory, key=lambda row: (row["role"], row["vaultPath"])),
        "privacy": "Row-level files are repo-local and Git-ignored; only this non-PII index is tracked.",
    }
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_path.write_text(json.dumps(index, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "inactiveManifest": str(inactive_path),
        "activeManifest": str(active_path),
        "index": str(index_path),
        "lanes": index["lanes"],
        "inactiveContactActions": dict(sorted(inactive_contact_counts.items())),
        "inactiveCourseActions": dict(sorted(inactive_course_counts.items())),
        "activeContactActions": dict(sorted(active_contact_counts.items())),
        "activeEnrollmentActions": dict(sorted(active_enrollment_counts.items())),
    }, indent=2))


if __name__ == "__main__":
    main()
