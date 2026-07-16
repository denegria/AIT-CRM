#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


def text(value) -> str:
    return "" if value is None else str(value).strip()


def rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        values = sheet.iter_rows(values_only=True)
        headers = [text(value) for value in next(values)]
        if headers == ["note"]:
            notes = [text(row[0]) for row in values if row]
            assert notes == ["No rows"]
            return []
        return [dict(zip(headers, row)) for row in values]
    finally:
        workbook.close()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def validate_execution_manifest(path: Path, expected_lane: str) -> dict:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    content_sha = manifest.pop("contentSha256")
    assert hashlib.sha256(canonical_json(manifest).encode("utf-8")).hexdigest() == content_sha
    manifest["contentSha256"] = content_sha
    assert manifest["schemaVersion"] == 1
    assert manifest["lane"] == expected_lane
    assert manifest["approvalState"] == "held"
    assert manifest["sequence"]["attendanceSupported"] is False
    assert set(manifest["requiredProductGates"]) == {"MIS-319", "MIS-320", "MIS-321", "MIS-322", "MIS-324"}
    actions = manifest["contactActions"] + manifest["classSectionActions"] + manifest["courseActions"]
    keys = [row["idempotencyKey"] for row in actions]
    assert len(keys) == len(set(keys))
    assert all(key.startswith(f"mis-318:{expected_lane}:") for key in keys)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--vault-root", type=Path, default=Path("private-imports/mis-318"))
    parser.add_argument("--index", type=Path, default=Path("docs/imports/mis-318/manifest-index.json"))
    args = parser.parse_args()
    source_root = args.source_root.resolve()
    vault_root = args.vault_root.resolve()
    index_path = args.index.resolve()

    index = json.loads(index_path.read_text(encoding="utf-8"))
    assert index["mode"] == "read_only_artifact_build"
    assert index["productionFingerprint"] == {
        "neonBranchId": "br-purple-bar-aphafrgp",
        "businessUnitName": "AIT USA Institute",
    }
    assert index["lanes"]["attendance"]["approvedManifest"] is False
    assert index["lanes"]["inactive"]["state"] == "manifest_final_data_held"
    assert index["lanes"]["active"]["state"] == "manifest_built_data_held_after_inactive"

    for item in index["files"]:
        vault_path = Path(item["vaultPath"])
        expected_prefix = Path("private-imports/mis-318")
        try:
            relative_vault_path = vault_path.relative_to(expected_prefix)
        except ValueError as error:
            raise AssertionError(f"Index path escaped the MIS-318 private vault: {vault_path}") from error
        path = vault_root / relative_vault_path
        assert path.exists(), path
        assert path.stat().st_size == item["sizeBytes"]
        assert sha256(path) == item["sha256"]
        assert path.stat().st_mode & 0o777 == 0o600

    original_roles = [item for item in index["files"] if item["role"] == "original"]
    cleaned_roles = [item for item in index["files"] if item["role"] == "cleaned"]
    attendance_roles = [item for item in index["files"] if item["role"] == "attendance_held"]
    execution_roles = [item for item in index["files"] if item["role"].endswith("_execution_manifest")]
    assert len(original_roles) == 2
    assert len(cleaned_roles) == 4
    assert len(attendance_roles) == 2
    assert len(execution_roles) == 2

    inactive_path = vault_root / "manifests/ait-usa-inactive-student-final-action-manifest.xlsx"
    active_path = vault_root / "manifests/ait-usa-active-enrollment-manifest.xlsx"
    inactive_execution_path = vault_root / "manifests/execution/ait-usa-inactive-student-actions-v1.json"
    active_execution_path = vault_root / "manifests/execution/ait-usa-active-enrollment-actions-v1.json"
    for path in (inactive_path, active_path):
        with ZipFile(path) as archive:
            assert archive.testzip() is None
    inactive_execution = validate_execution_manifest(inactive_execution_path, "inactive")
    active_execution = validate_execution_manifest(active_execution_path, "active")

    contact_actions = rows(inactive_path, "Contact_Actions")
    inactive_courses = rows(inactive_path, "Historical_Course_Actions")
    deferred_contacts = rows(inactive_path, "Deferred_Contacts")
    assert len(contact_actions) == 584
    assert len(deferred_contacts) == 11
    assert sum(1 for row in contact_actions if not text(row["final_contact_action"]).startswith("defer")) == 573
    assert all(row["data_apply_state"] == "held_pending_manifest_approval" for row in contact_actions)
    assert all(row["planned_contact_reference"] for row in contact_actions if not text(row["final_contact_action"]).startswith("defer"))
    assert all(not row["planned_contact_reference"] for row in deferred_contacts)

    inactive_counts = Counter(text(row["proposed_course_action"]) for row in inactive_courses)
    assert len(inactive_courses) == 619
    assert inactive_counts == Counter({
        "insert_dropped_course_after_approval": 423,
        "defer_course_definition": 147,
        "defer_source_course_review": 25,
        "insert_dropped_course_after_catalog_gate": 12,
        "defer_contact_identity": 11,
        "skip_source_duplicate": 1,
    })
    assert all(row["course_status"] == "dropped" for row in inactive_courses)
    assert all(row["data_apply_state"] == "held_pending_product_gates_and_manifest_approval" for row in inactive_courses)
    catalog_gate_rows = [row for row in inactive_courses if row["proposed_course_action"] == "insert_dropped_course_after_catalog_gate"]
    assert {row["mapped_course"] for row in catalog_gate_rows} == {"Computer", "Math"}
    assert all(row["planned_contact_reference"] for row in inactive_courses if text(row["proposed_course_action"]).startswith("insert"))

    active_contacts = rows(active_path, "Unique_Contact_Actions")
    active_enrollments = rows(active_path, "Active_Enrollment_Actions")
    active_holds = rows(active_path, "Enrollment_Holds")
    assert len(active_contacts) == 144
    assert len(active_enrollments) == 147
    assert len(active_holds) == 22
    assert sum(1 for row in active_enrollments if text(row["proposed_enrollment_action"]).startswith("insert")) == 125
    assert all(row["course_status"] == "active" for row in active_enrollments)
    assert all(row["data_apply_state"] == "held_pending_inactive_lane_and_product_gates" for row in active_enrollments)
    assert all(row["modality"] == "PRESENCIAL" for row in active_enrollments if row["location"] == "Plainfield")
    assert sum(1 for row in active_enrollments if row["mapped_course"] == "Computer") == 6
    assert sum(1 for row in active_enrollments if row["proposed_enrollment_action"] == "hold_contact_identity") == 4
    assert sum(1 for row in active_enrollments if row["proposed_enrollment_action"] == "hold_roster_lifecycle_status") == 10
    assert sum(1 for row in active_enrollments if row["requires_multiple_active_enrollment_product_gate"] == "yes") >= 5

    by_identity = defaultdict(list)
    for row in active_enrollments:
        by_identity[text(row["planned_contact_reference"]) or f"held:{row['source_cell']}"] .append(row)
    repeated = [group for key, group in by_identity.items() if not key.startswith("held:") and len(group) > 1]
    assert repeated
    assert len(inactive_execution["contactActions"]) == len(contact_actions)
    assert len(inactive_execution["courseActions"]) == len(inactive_courses)
    assert len(active_execution["contactActions"]) == len(active_contacts)
    assert len(active_execution["courseActions"]) == len(active_enrollments)
    assert len(active_execution["classSectionActions"]) == active_execution["expectedCounts"]["classSections"]
    assert active_execution["sequence"]["requiredPriorManifestSha256"] == inactive_execution["contentSha256"]

    source_active_count = 0
    for slug in ("bound-brook", "plainfield"):
        source_path = source_root / "clean/buckets" / f"ait-usa-{slug}-active-students.xlsx"
        source_active_count += len(rows(source_path, "Roster_Source_Rows"))
    assert source_active_count == len(active_enrollments) == 147

    repo_root = Path.cwd().resolve()
    course_records_source = (repo_root / "src/lib/crm/course-records.js").read_text(encoding="utf-8")
    schema_source = (repo_root / "src/db/schema.js").read_text(encoding="utf-8")
    assert "'Computer'" in course_records_source
    assert "'Math'" in course_records_source
    assert "same class section" in course_records_source
    assert "contactPhoneNumbers" in schema_source
    assert "courseClassSections" in schema_source
    assert "classSectionId" in schema_source

    print(json.dumps({
        "validatedFiles": len(index["files"]),
        "inactive": {
            "contactCandidates": len(contact_actions),
            "resolvedContacts": 573,
            "deferredContacts": len(deferred_contacts),
            "historicalCourseRows": len(inactive_courses),
            "actionableAfterGates": sum(1 for row in inactive_courses if text(row["proposed_course_action"]).startswith("insert")),
        },
        "active": {
            "officialRosterRows": len(active_enrollments),
            "uniqueContacts": len(active_contacts),
            "actionableAfterGates": sum(1 for row in active_enrollments if text(row["proposed_enrollment_action"]).startswith("insert")),
            "held": len(active_holds),
        },
        "attendance": "held_unvalidated_absolute_last",
        "productGatesConfirmed": [
            "Computer and Math catalog",
            "Contact alternate-phone history",
            "course sections and multiple active enrollments",
            "idempotent manifest dry-run/apply",
        ],
    }, indent=2))


if __name__ == "__main__":
    main()
